from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── 색상 상수 ──
DARK_BLUE    = "1A3A5C"
MID_BLUE     = "2980B9"
LIGHT_BLUE   = "EAF3FB"
ORANGE       = "E67E22"
LIGHT_ORANGE = "FEF5E7"
GREEN        = "27AE60"
LIGHT_GREEN  = "EAFAF1"
RED          = "C0392B"
LIGHT_RED    = "FDECEA"
GRAY         = "7F8C8D"
LIGHT_GRAY   = "F4F7FB"
WHITE        = "FFFFFF"

def hdr_font(size=11):
    return Font(name="Arial", size=size, bold=True, color=WHITE)

def body_font(size=10, bold=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def blue_font(size=10):
    return Font(name="Arial", size=size, color="0000FF")

def black_font(size=10, bold=False):
    return Font(name="Arial", size=size, bold=bold, color="000000")

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="D0DAE8")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center")

def left_align():
    return Alignment(horizontal="left", vertical="center")


# ═══════════════════════════════════════════════════════
# SHEET 1 — 월간 단가 입력
# ═══════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "월간단가입력"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "C4"

# 제목
ws1.merge_cells("A1:N1")
ws1["A1"] = "PP / PE 레진 월간 단가 추적 시트"
ws1["A1"].font = Font(name="Arial", size=16, bold=True, color=WHITE)
ws1["A1"].fill = fill(DARK_BLUE)
ws1["A1"].alignment = center()
ws1.row_dimensions[1].height = 38

ws1.merge_cells("A2:N2")
ws1["A2"] = "파란색(Blue) = 직접 입력  |  검은색(Black) = 자동 계산 수식  |  노란색 배경 = 주의 필요 항목"
ws1["A2"].font = Font(name="Arial", size=9, color=GRAY)
ws1["A2"].fill = fill(LIGHT_BLUE)
ws1["A2"].alignment = left_align()
ws1.row_dimensions[2].height = 18

# 헤더
headers = [
    "연도", "월",
    "PP Homo\n(USD/MT)", "PP Copo\n(USD/MT)",
    "HDPE\n(USD/MT)", "LDPE\n(USD/MT)", "LLDPE\n(USD/MT)",
    "나프타\n(USD/MT)", "Brent\n(USD/bbl)", "USD/KRW\n(원)",
    "PP-나프타\n스프레드", "PP Homo\n(원/MT)",
    "전월대비\nPP변동", "비고"
]
ws1.row_dimensions[3].height = 36
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=3, column=col, value=h)
    c.font = hdr_font(size=9)
    c.fill = fill(DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border()

# 열 너비
col_widths = [8, 6, 13, 13, 13, 13, 13, 13, 12, 12, 14, 15, 13, 22]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# 데이터 (2024-01 ~ 2025-06)
data_rows = [
    (2024,  1,  990, 1050,  955, 1035,  930,  645, 78.0, 1320),
    (2024,  2, 1010, 1070,  975, 1055,  950,  660, 80.0, 1330),
    (2024,  3, 1020, 1080,  985, 1065,  960,  670, 83.0, 1340),
    (2024,  4, 1040, 1100, 1005, 1085,  980,  685, 88.0, 1365),
    (2024,  5, 1040, 1100, 1005, 1085,  980,  685, 86.0, 1365),
    (2024,  6, 1020, 1080,  988, 1068,  963,  665, 84.0, 1370),
    (2024,  7, 1020, 1080,  988, 1068,  963,  665, 82.0, 1375),
    (2024,  8,  970, 1030,  940, 1020,  915,  620, 76.0, 1330),
    (2024,  9,  970, 1030,  940, 1020,  915,  620, 74.0, 1330),
    (2024, 10,  950, 1010,  920, 1000,  895,  600, 72.0, 1355),
    (2024, 11,  950, 1010,  920, 1000,  895,  600, 72.0, 1355),
    (2024, 12,  960, 1020,  930, 1010,  905,  610, 73.0, 1360),
    (2025,  1,  980, 1040,  950, 1030,  930,  630, 76.0, 1430),
    (2025,  2,  960, 1020,  930, 1010,  910,  610, 72.0, 1460),
    (2025,  3,  960, 1020,  930, 1010,  910,  610, 68.0, 1460),
    (2025,  4,  970, 1030,  940, 1020,  920,  615, 69.0, 1450),
    (2025,  5, 1000, 1060,  970, 1050,  950,  620, 67.0, 1420),
    (2025,  6, 1010, 1070,  990, 1060,  970,  630, 74.0, 1380),
]

for i, row in enumerate(data_rows):
    r = i + 4
    yr, mo, pp, pc, hd, ld, ll, nap, br, fx = row
    alt_fill = fill(LIGHT_BLUE) if i % 2 == 0 else fill(WHITE)
    ws1.row_dimensions[r].height = 20

    def set_cell(col, val, fnt, fmt=None):
        c = ws1.cell(row=r, column=col, value=val)
        c.font = fnt
        c.fill = alt_fill
        c.border = thin_border()
        c.alignment = center()
        if fmt:
            c.number_format = fmt

    set_cell(1,  yr,  blue_font(), "#,##0")
    set_cell(2,  mo,  blue_font(), "00")
    set_cell(3,  pp,  blue_font(), "#,##0")
    set_cell(4,  pc,  blue_font(), "#,##0")
    set_cell(5,  hd,  blue_font(), "#,##0")
    set_cell(6,  ld,  blue_font(), "#,##0")
    set_cell(7,  ll,  blue_font(), "#,##0")
    set_cell(8,  nap, blue_font(), "#,##0")
    set_cell(9,  br,  blue_font(), "#,##0.0")
    set_cell(10, fx,  blue_font(), "#,##0")

    # 스프레드 수식
    cs = ws1.cell(row=r, column=11, value="=C{0}-H{0}".format(r))
    cs.font = black_font()
    cs.fill = alt_fill
    cs.border = thin_border()
    cs.alignment = center()
    cs.number_format = "#,##0"

    # 원화 환산 수식
    ck = ws1.cell(row=r, column=12, value="=C{0}*J{0}".format(r))
    ck.font = black_font()
    ck.fill = alt_fill
    ck.border = thin_border()
    ck.alignment = center()
    ck.number_format = "#,##0"

    # 전월 대비 변동
    if i == 0:
        cv = ws1.cell(row=r, column=13, value="-")
        cv.font = black_font()
    else:
        cv = ws1.cell(row=r, column=13, value="=C{0}-C{1}".format(r, r-1))
        cv.font = black_font()
        cv.number_format = "+#,##0;-#,##0;\"-\""
    cv.fill = alt_fill
    cv.border = thin_border()
    cv.alignment = center()

    # 비고
    cn = ws1.cell(row=r, column=14, value="")
    cn.fill = alt_fill
    cn.border = thin_border()
    cn.font = blue_font()

# 빈 입력 행 (6개)
last_data_r = len(data_rows) + 4
for i in range(6):
    r = last_data_r + i
    ws1.row_dimensions[r].height = 22
    for col in range(1, 15):
        c = ws1.cell(row=r, column=col, value="")
        c.fill = PatternFill("solid", fgColor="EAF3FB")
        c.border = thin_border()
        c.alignment = center()
        if col in (3, 4, 5, 6, 7, 8, 9, 10):
            c.font = blue_font()
        elif col == 11:
            c.value = "=IF(C{0}=\"\",\"\",C{0}-H{0})".format(r)
            c.font = black_font()
            c.number_format = "#,##0"
        elif col == 12:
            c.value = "=IF(C{0}=\"\",\"\",C{0}*J{0})".format(r)
            c.font = black_font()
            c.number_format = "#,##0"
        elif col == 13 and i > 0:
            c.value = "=IF(C{0}=\"\",\"\",C{0}-C{1})".format(r, r-1)
            c.font = black_font()
            c.number_format = "+#,##0;-#,##0;\"-\""
        else:
            c.font = black_font()


# ═══════════════════════════════════════════════════════
# SHEET 2 — 공급사별 오퍼 비교
# ═══════════════════════════════════════════════════════
ws2 = wb.create_sheet("공급사오퍼비교")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:L1")
ws2["A1"] = "공급사별 오퍼 비교표"
ws2["A1"].font = Font(name="Arial", size=15, bold=True, color=WHITE)
ws2["A1"].fill = fill(DARK_BLUE)
ws2["A1"].alignment = center()
ws2.row_dimensions[1].height = 36

ws2.merge_cells("A2:L2")
ws2["A2"] = "ICIS 기준가 대비 프리미엄/디스카운트 자동 계산  |  파란색 셀에 오퍼가 입력"
ws2["A2"].font = Font(name="Arial", size=9, color=GRAY)
ws2["A2"].fill = fill(LIGHT_BLUE)
ws2["A2"].alignment = left_align()

# ICIS 기준가
ws2.merge_cells("A3:B3")
ws2["A3"] = "ICIS 기준가 (USD/MT)"
ws2["A3"].font = Font(name="Arial", size=10, bold=True, color=DARK_BLUE)
ws2["A3"].alignment = left_align()

basis_labels = ["PP Homo", "PP Copo", "HDPE", "LDPE", "LLDPE"]
basis_vals   = [1010, 1070, 990, 1060, 970]
for j, (lab, val) in enumerate(zip(basis_labels, basis_vals)):
    col = j + 3
    cl = ws2.cell(row=3, column=col, value=lab)
    cl.font = Font(name="Arial", size=9, bold=True, color=WHITE)
    cl.fill = fill(MID_BLUE)
    cl.alignment = center()
    cv = ws2.cell(row=4, column=col, value=val)
    cv.font = blue_font()
    cv.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    cv.alignment = center()
    cv.number_format = "#,##0"
    cv.border = thin_border()

# 헤더
offer_headers = [
    "공급사", "원산지",
    "PP Homo\n오퍼", "PP Copo\n오퍼", "HDPE\n오퍼", "LDPE\n오퍼", "LLDPE\n오퍼",
    "PP 프리미엄\n(vs ICIS)", "결제조건", "리드타임", "MOQ(MT)", "추천"
]
ws2.row_dimensions[5].height = 36
for col, h in enumerate(offer_headers, 1):
    c = ws2.cell(row=5, column=col, value=h)
    c.font = hdr_font(size=9)
    c.fill = fill(DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border()

suppliers = [
    ("LG화학",      "한국",    1015, 1075,  993, 1063,  973, "T/T 30일", "2주", 20),
    ("롯데케미칼",  "한국",    1008, 1068,  988, 1058,  968, "T/T 30일", "2주", 20),
    ("한화솔루션",  "한국",    1020, 1080, 1000, 1070,  980, "T/T 30일", "3주", 25),
    ("SABIC",       "사우디",   995, 1055,  975, 1045,  955, "L/C 60일", "5주", 50),
    ("Borouge",     "UAE",      990, 1050,  970, 1040,  950, "L/C 60일", "5주", 50),
    ("QAPCO",       "카타르",   985, 1045,  965, 1035,  945, "L/C 60일", "6주", 50),
    ("PTT GC",      "태국",    1000, 1060,  980, 1050,  960, "T/T 45일", "3주", 30),
]

for i, (sup, ori, pp, pc, hd, ld, ll, pay, lead, moq) in enumerate(suppliers):
    r = i + 6
    ws2.row_dimensions[r].height = 20
    alt_fill = fill(LIGHT_BLUE) if i % 2 == 0 else fill(WHITE)

    def sc(col, val, fnt=None, fmt=None):
        c = ws2.cell(row=r, column=col, value=val)
        c.font = fnt or body_font()
        c.fill = alt_fill
        c.border = thin_border()
        c.alignment = center()
        if fmt:
            c.number_format = fmt

    sc(1,  sup, body_font(bold=True))
    sc(2,  ori)
    sc(3,  pp,  blue_font(), "#,##0")
    sc(4,  pc,  blue_font(), "#,##0")
    sc(5,  hd,  blue_font(), "#,##0")
    sc(6,  ld,  blue_font(), "#,##0")
    sc(7,  ll,  blue_font(), "#,##0")

    cp = ws2.cell(row=r, column=8, value="=C{0}-C4".format(r))
    cp.font = black_font()
    cp.fill = alt_fill
    cp.border = thin_border()
    cp.alignment = center()
    cp.number_format = "+#,##0;-#,##0;\"-\""

    sc(9,  pay)
    sc(10, lead)
    sc(11, moq, fmt="#,##0")

    cr = ws2.cell(row=r, column=12, value='=IF(H{0}<=0,"★ 추천",IF(H{0}<=20,"검토","-"))'.format(r))
    cr.font = black_font()
    cr.fill = alt_fill
    cr.border = thin_border()
    cr.alignment = center()

ws2_widths = [14, 10, 12, 12, 12, 12, 12, 14, 12, 10, 10, 10]
for i, w in enumerate(ws2_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════════════
# SHEET 3 — 단가 시뮬레이터
# ═══════════════════════════════════════════════════════
ws3 = wb.create_sheet("단가시뮬레이터")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:H1")
ws3["A1"] = "PP / PE 레진 단가 시뮬레이터  (유가 · 환율 · 나프타 변동 영향)"
ws3["A1"].font = Font(name="Arial", size=14, bold=True, color=WHITE)
ws3["A1"].fill = fill(DARK_BLUE)
ws3["A1"].alignment = center()
ws3.row_dimensions[1].height = 36

ws3.merge_cells("A3:C3")
ws3["A3"] = "▶ 입력 가정 (파란색 셀 수정)"
ws3["A3"].font = Font(name="Arial", size=11, bold=True, color=DARK_BLUE)

assumptions = [
    ("Brent 유가 (USD/bbl)",         74,    "#,##0.0", "현재 시장가"),
    ("나프타 CFR Japan (USD/MT)",    630,    "#,##0",   "Platts 기준"),
    ("USD/KRW 환율 (원)",           1380,   "#,##0",   "시장 환율"),
    ("PP-나프타 배율",               1.55,   "0.00",    "공급사 협의 배율"),
    ("고정 마진 (USD/MT)",             85,   "#,##0",   "계약 고정분"),
    ("물류비 (USD/MT)",                20,   "#,##0",   "CFR 기준"),
]

for j, (label, val, fmt, note) in enumerate(assumptions):
    r = j + 5
    ws3.row_dimensions[r].height = 22
    cl = ws3.cell(row=r, column=1, value=label)
    cl.font = body_font(bold=True)
    cl.alignment = left_align()
    cv = ws3.cell(row=r, column=2, value=val)
    cv.font = blue_font()
    cv.fill = PatternFill("solid", fgColor="EAF3FB")
    cv.border = thin_border()
    cv.alignment = center()
    cv.number_format = fmt
    cn = ws3.cell(row=r, column=3, value=note)
    cn.font = Font(name="Arial", size=9, color=GRAY, italic=True)
    cn.alignment = left_align()

# 전월 비교 기준가
ws3["E3"] = "전월 기준가 (비교용)"
ws3["E3"].font = Font(name="Arial", size=10, bold=True, color=DARK_BLUE)
ws3["E3"].fill = PatternFill("solid", fgColor=LIGHT_ORANGE)

for col, label, val in [(5, "PP Homo", 980), (6, "HDPE", 960)]:
    ws3.cell(row=4, column=col, value=label).font = Font(name="Arial", size=9, bold=True, color=GRAY)
    ws3.cell(row=4, column=col).alignment = center()
    cv = ws3.cell(row=5, column=col, value=val)
    cv.font = blue_font()
    cv.border = thin_border()
    cv.alignment = center()
    cv.number_format = "#,##0"

ws3.merge_cells("A12:H12")
ws3["A12"] = "▶ 산출 결과 (자동 계산)"
ws3["A12"].font = Font(name="Arial", size=11, bold=True, color=DARK_BLUE)
ws3.row_dimensions[12].height = 24

result_headers = ["품목", "산출 공식 개요", "USD 단가\n(USD/MT)", "원화 환산\n(원/MT)",
                  "전월 USD\n차이", "전월 원화\n차이", "등락률"]
ws3.row_dimensions[13].height = 36
for col, h in enumerate(result_headers, 1):
    c = ws3.cell(row=13, column=col, value=h)
    c.font = hdr_font(size=9)
    c.fill = fill(DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border()

results = [
    ("PP Homo (나프타 연동)",        "=B6*B8+B9+B10",            "E5", True),
    ("PP Copo (+$60 프리미엄)",      "=B6*B8+B9+B10+60",          None, False),
    ("HDPE (에틸렌 연동 근사)",      "=(B6+220)*1.28+B9+B10",    "F5", True),
    ("LDPE (+$70 프리미엄)",         "=(B6+220)*1.28+B9+B10+70", None, False),
    ("LLDPE C4",                    "=(B6+220)*1.25+B9+B10",    None, False),
]

for j, (name, formula, prev_ref, show_diff) in enumerate(results):
    r = j + 14
    ws3.row_dimensions[r].height = 22
    alt_fill = fill(LIGHT_BLUE) if j % 2 == 0 else fill(WHITE)

    cl = ws3.cell(row=r, column=1, value=name)
    cl.font = body_font(bold=True)
    cl.fill = alt_fill
    cl.border = thin_border()
    cl.alignment = left_align()

    cf = ws3.cell(row=r, column=2, value=formula.replace("=",""))
    cf.font = Font(name="Arial", size=8, color=GRAY, italic=True)
    cf.fill = alt_fill
    cf.border = thin_border()
    cf.alignment = left_align()

    cu = ws3.cell(row=r, column=3, value=formula)
    cu.font = black_font(bold=True)
    cu.fill = alt_fill
    cu.border = thin_border()
    cu.alignment = center()
    cu.number_format = "#,##0"

    ck = ws3.cell(row=r, column=4, value="=C{0}*B7".format(r))
    ck.font = black_font()
    ck.fill = alt_fill
    ck.border = thin_border()
    ck.alignment = center()
    ck.number_format = "#,##0"

    if show_diff and prev_ref:
        cd = ws3.cell(row=r, column=5, value="=C{0}-{1}".format(r, prev_ref))
        cd.number_format = "+#,##0;-#,##0;\"-\""
        ckd = ws3.cell(row=r, column=6, value="=D{0}-{1}*B7".format(r, prev_ref))
        ckd.number_format = "+#,##0;-#,##0;\"-\""
        cp  = ws3.cell(row=r, column=7, value='=IF({0}=0,"-",(C{1}-{0})/{0})'.format(prev_ref, r))
        cp.number_format = "0.0%;-0.0%;\"-\""
    else:
        cd  = ws3.cell(row=r, column=5, value="-")
        ckd = ws3.cell(row=r, column=6, value="-")
        cp  = ws3.cell(row=r, column=7, value="-")

    for c in (cd, ckd, cp):
        c.font = black_font()
        c.fill = alt_fill
        c.border = thin_border()
        c.alignment = center()

ws3_widths = [28, 30, 14, 16, 14, 16, 10, 5]
for i, w in enumerate(ws3_widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════════════
# SHEET 4 — KPI 요약 대시보드
# ═══════════════════════════════════════════════════════
ws4 = wb.create_sheet("KPI요약")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:I1")
ws4["A1"] = "PP / PE 레진 구매 KPI 월간 요약"
ws4["A1"].font = Font(name="Arial", size=15, bold=True, color=WHITE)
ws4["A1"].fill = fill(DARK_BLUE)
ws4["A1"].alignment = center()
ws4.row_dimensions[1].height = 36

kpi_items = [
    (3, 1,  MID_BLUE,   "PP Homo 구매 단가",       "=월간단가입력!C21", "#,##0",          "USD/MT"),
    (3, 4,  ORANGE,     "전월 대비 PP 변동",        "=월간단가입력!M21", "+#,##0;-#,##0;0","USD/MT"),
    (3, 7,  GREEN,      "PP-나프타 스프레드",       "=월간단가입력!K21", "#,##0",          "USD/MT"),
    (7, 1,  RED,        "PP Homo 원화 환산 단가",   "=월간단가입력!L21", "#,##0",          "원/MT"),
    (7, 4,  MID_BLUE,   "USD/KRW 환율",             "=월간단가입력!J21", "#,##0",          "원"),
    (7, 7,  GRAY,       "Brent 유가",               "=월간단가입력!I21", "#,##0.0",        "USD/bbl"),
]

for row_start, col_start, color, title, formula, fmt, unit in kpi_items:
    for row_offset in range(3):
        ws4.merge_cells(
            start_row=row_start + row_offset,
            start_column=col_start,
            end_row=row_start + row_offset,
            end_column=col_start + 1
        )
    ct = ws4.cell(row=row_start,     column=col_start, value=title)
    cv = ws4.cell(row=row_start + 1, column=col_start, value=formula)
    cu = ws4.cell(row=row_start + 2, column=col_start, value=unit)

    ct.font = Font(name="Arial", size=9, bold=True, color=WHITE)
    ct.fill = PatternFill("solid", fgColor=color)
    ct.alignment = center()

    cv.font = Font(name="Arial", size=18, bold=True, color=color)
    cv.fill = fill(WHITE)
    cv.alignment = center()
    cv.number_format = fmt
    cv.border = thin_border()

    cu.font = Font(name="Arial", size=9, color=GRAY)
    cu.fill = fill(LIGHT_GRAY)
    cu.alignment = center()

for col_letter, w in [("A",18),("B",18),("C",3),("D",18),("E",18),("F",3),("G",18),("H",18)]:
    ws4.column_dimensions[col_letter].width = w

# 스프레드 신호등 테이블
ws4["A11"] = "▶ PP-나프타 스프레드 신호등"
ws4["A11"].font = Font(name="Arial", size=11, bold=True, color=DARK_BLUE)

signal_headers = ["스프레드 구간", "신호", "해석", "구매 액션"]
for j, h in enumerate(signal_headers, 1):
    c = ws4.cell(row=12, column=j, value=h)
    c.font = hdr_font(size=9)
    c.fill = fill(DARK_BLUE)
    c.alignment = center()
    c.border = thin_border()

signals = [
    ("$400 이상",  "GREEN",  "공급 여유 / 생산자 감산 없음",   "스팟 구매 검토, 재고 절감",   LIGHT_GREEN),
    ("$300~$400", "YELLOW", "정상 마진 범위",                  "계획 구매 유지",               "FFFFF0"),
    ("$200~$300", "ORANGE", "마진 압박 / 감산 가능성",          "장기 계약 비중 확대",          LIGHT_ORANGE),
    ("$200 미만",  "RED",    "감산 임박 / 공급 타이트 예상",    "긴급 재고 확보 필요",          LIGHT_RED),
]
for i, (span, sig, desc, action, clr) in enumerate(signals):
    r = i + 13
    ws4.row_dimensions[r].height = 22
    for j, val in enumerate([span, sig, desc, action], 1):
        c = ws4.cell(row=r, column=j, value=val)
        c.font = body_font()
        c.fill = PatternFill("solid", fgColor=clr)
        c.border = thin_border()
        c.alignment = center()

for j, w in [(1, 14), (2, 12), (3, 30), (4, 25)]:
    ws4.column_dimensions[get_column_letter(j)].width = w


# ═══════════════════════════════════════════════════════
# 저장
# ═══════════════════════════════════════════════════════
out_path = r"C:\Temp\ai-workshop\PP_PE_레진_단가추적.xlsx"
wb.save(out_path)
print("저장 완료:", out_path)
