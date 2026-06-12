from flask import Flask, jsonify, request, send_from_directory, send_file
import json, os, smtplib, io
from dotenv import load_dotenv
load_dotenv()
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import date, datetime

app = Flask(__name__)
BASE = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.join(BASE, 'data', 'data.json')


def load_data():
    with open(DATA_FILE, 'r', encoding='utf-8-sig') as f:
        return json.load(f)


def save_data(data):
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ── Static files ──────────────────────────────────────
@app.route('/')
def index():
    return send_from_directory(BASE, 'index.html')

@app.route('/<path:path>')
def static_files(path):
    return send_from_directory(BASE, path)


# ── API: data ─────────────────────────────────────────
@app.route('/api/data', methods=['GET'])
def api_get():
    return jsonify(load_data())

VALID_STAGES   = {'네고중','계약완료','발주대기','선적완료','항해중','통관중','배차완료'}
VALID_NEGO_ST  = {'진행중','완료'}
VALID_CUST_ST  = {'대기중','진행중','완료'}
VALID_DISP_ST  = {'대기중','진행중','완료'}

def validate_orders(orders):
    errors = []
    for o in orders:
        oid = o.get('id', '?')
        if o.get('stage') and o['stage'] not in VALID_STAGES:
            errors.append(f"{oid}: stage '{o['stage']}' 허용값 아님")
        n = o.get('negotiation') or {}
        if n.get('status') and n['status'] not in VALID_NEGO_ST:
            errors.append(f"{oid}: negotiation.status '{n['status']}' 허용값 아님")
        c = o.get('customs') or {}
        if c.get('status') and c['status'] not in VALID_CUST_ST:
            errors.append(f"{oid}: customs.status '{c['status']}' 허용값 아님")
        d = o.get('dispatch') or {}
        if d.get('status') and d['status'] not in VALID_DISP_ST:
            errors.append(f"{oid}: dispatch.status '{d['status']}' 허용값 아님")
    return errors

@app.route('/api/data', methods=['POST'])
def api_save():
    data = request.get_json(force=True)
    if not data:
        return jsonify({'ok': False, 'error': '데이터 없음'}), 400
    errors = validate_orders(data.get('orders', []))
    if errors:
        return jsonify({'ok': False, 'error': '검증 실패', 'details': errors}), 400
    save_data(data)
    return jsonify({'ok': True})


# ── API: Excel export ─────────────────────────────────
@app.route('/api/export')
def api_export():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'error': 'openpyxl 미설치. pip install openpyxl'}), 500

    data = load_data()
    wb = openpyxl.Workbook()

    # ── 공통 스타일 ──
    header_font  = Font(bold=True, color='FFFFFF', size=10)
    header_fill  = PatternFill('solid', fgColor='1E40AF')
    center       = Alignment(horizontal='center', vertical='center')
    left         = Alignment(horizontal='left',   vertical='center')
    thin         = Side(style='thin', color='CBD5E1')
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, headers, col_widths):
        ws.append(headers)
        for i, w in enumerate(col_widths, 1):
            c = ws.cell(1, i)
            c.font, c.fill, c.alignment, c.border = header_font, header_fill, center, border
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 22

    def style_row(ws, row_num, n_cols):
        fill = PatternFill('solid', fgColor='F8FAFC') if row_num % 2 == 0 else None
        for i in range(1, n_cols + 1):
            c = ws.cell(row_num, i)
            c.border = border
            c.alignment = left
            if fill:
                c.fill = fill

    # ── 시트 1: 오더 현황 ──
    ws1 = wb.active
    ws1.title = '오더 현황'
    ws1.freeze_panes = 'A2'
    headers1 = ['오더번호','품목','등급','공급사','원산지','수량(MT)','단계',
                 '오퍼가','카운터가','계약가','시장가','협상상태','네고일',
                 '계약번호','계약일','결제조건','인도조건','납기월',
                 '선박명','B/L No.','ETD','ETA','선적항','도착항',
                 '입항일','신고일','통관완료일','통관상태',
                 '배차상태','배차일','목적지']
    widths1 = [14,10,10,14,10,10,10,
               10,10,10,10,10,12,
               16,12,12,12,10,
               14,16,12,12,12,12,
               12,12,12,10,
               10,12,14]
    style_header(ws1, headers1, widths1)

    for i, o in enumerate(data.get('orders', []), 2):
        n = o.get('negotiation', {})
        c = o.get('contract', {})
        s = o.get('shipment', {})
        cu = o.get('customs', {})
        d = o.get('dispatch', {})
        ws1.append([
            o.get('id'), o.get('product'), o.get('grade'), o.get('supplier'), o.get('origin'),
            o.get('quantity'), o.get('stage'),
            n.get('offerPrice'), n.get('counterPrice'), n.get('contractPrice'), n.get('marketRef'),
            n.get('status'), n.get('date'),
            c.get('contractNo'), c.get('signDate'), c.get('paymentTerms'), c.get('incoterms'), c.get('deliveryMonth'),
            s.get('vessel'), s.get('blNo'), s.get('etd'), s.get('eta'), s.get('portOfLoading'), s.get('portOfDischarge'),
            cu.get('arrivalDate'), cu.get('declarationDate'), cu.get('clearanceDate'), cu.get('status'),
            d.get('status'), d.get('date'), d.get('destination')
        ])
        style_row(ws1, i, len(headers1))

    # ── 시트 2: 재고 현황 ──
    ws2 = wb.create_sheet('재고 현황')
    ws2.freeze_panes = 'A2'
    headers2 = ['품목','창고','현재고(MT)','안전재고(MT)','입고예정(MT)','출고예정(MT)',
                 '예상잔고(MT)','일평균사용(MT)','재고일수(DOI)','안전재고 대비']
    style_header(ws2, headers2, [12,14,12,12,12,12,12,14,12,14])

    for i, inv in enumerate(data.get('inventory', []), 2):
        cur  = inv.get('currentStock', 0)
        safe = inv.get('safetyStock', 0)
        inc  = inv.get('incomingStock', 0)
        out  = inv.get('outgoingPlan', 0)
        avg  = inv.get('avgDailyUsage', 0)
        proj = cur + inc - out
        doi  = round(cur / avg, 1) if avg > 0 else '∞'
        diff = cur - safe
        ws2.append([inv.get('product'), inv.get('location'), cur, safe, inc, out, proj, avg, doi,
                    f'+{diff}MT' if diff >= 0 else f'{diff}MT'])
        style_row(ws2, i, len(headers2))
        # 안전재고 미달 시 빨간 글씨
        if cur < safe:
            for col in [3, 10]:
                ws2.cell(i, col).font = Font(bold=True, color='DC2626')

    # ── 시트 3: 시장가 이력 ──
    ws3 = wb.create_sheet('시장가 이력')
    ws3.freeze_panes = 'A2'
    headers3 = ['월','PP Homo','PP Copo','HDPE','LDPE','LLDPE','나프타','Brent($/bbl)']
    style_header(ws3, headers3, [10,10,10,10,10,10,10,12])

    for i, mp in enumerate(data.get('marketPrices', []), 2):
        ws3.append([mp.get('month'), mp.get('ppHomo'), mp.get('ppCopo'), mp.get('hdpe'),
                    mp.get('ldpe'), mp.get('lldpe'), mp.get('naphtha'), mp.get('brent')])
        style_row(ws3, i, len(headers3))

    # ── 파일 저장 후 스트리밍 ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'resin_tracker_{date.today().strftime("%Y%m%d")}.xlsx'
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


# ── API: Email alert ──────────────────────────────────
@app.route('/api/alert', methods=['POST'])
def api_alert():
    body = request.get_json(force=True) or {}
    to_addr   = body.get('to', os.environ.get('EMAIL_TO', ''))
    from_addr = os.environ.get('EMAIL_FROM', '')
    password  = os.environ.get('EMAIL_PASSWORD', '')
    smtp_host = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
    smtp_port = int(os.environ.get('SMTP_PORT', 587))

    if not all([to_addr, from_addr, password]):
        return jsonify({'ok': False, 'error': '이메일 설정 없음. EMAIL_FROM, EMAIL_PASSWORD, EMAIL_TO 환경변수를 설정하세요.'}), 400

    data   = load_data()
    today  = date.today()
    alerts = []

    # 안전재고 미달
    for inv in data.get('inventory', []):
        if inv.get('currentStock', 0) < inv.get('safetyStock', 0):
            alerts.append(f"⚠️ {inv['product']} 현재고 {inv['currentStock']}MT — 안전재고({inv['safetyStock']}MT) 미달")

    # ETA 임박 오더 (항해중)
    for o in data.get('orders', []):
        eta = o.get('shipment', {}).get('eta', '')
        if eta and o.get('stage') == '항해중':
            try:
                days = (datetime.strptime(eta, '%Y-%m-%d').date() - today).days
                if 0 <= days <= 14:
                    alerts.append(f"🚢 {o['id']} ({o['product']}) ETA {eta} — D-{days}")
            except Exception:
                pass

    # 통관 진행 중
    for o in data.get('orders', []):
        if o.get('stage') == '통관중':
            alerts.append(f"🏛 {o['id']} ({o['product']}) 통관 진행 중 — 완료 확인 필요")

    if not alerts:
        return jsonify({'ok': True, 'message': '알림 조건 없음 (긴급 건 없음)'})

    # 메일 작성
    subject = f'[레진 구매] 긴급 알림 {len(alerts)}건 — {today.strftime("%Y-%m-%d")}'
    html = f"""
<html><body style="font-family:sans-serif;max-width:600px;margin:0 auto">
<h2 style="color:#1e40af">🏭 PP/PE 레진 구매 알림</h2>
<p style="color:#475569">{today.strftime('%Y년 %m월 %d일')} 기준 긴급 처리 필요 항목입니다.</p>
<ul style="line-height:2">
{''.join(f'<li>{a}</li>' for a in alerts)}
</ul>
<hr>
<p style="color:#94a3b8;font-size:12px">PP/PE 레진 구매 관리 시스템 자동 발송</p>
</body></html>"""

    msg = MIMEMultipart('alternative')
    msg['Subject'], msg['From'], msg['To'] = subject, from_addr, to_addr
    msg.attach(MIMEText(html, 'html', 'utf-8'))

    try:
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.ehlo()
            server.starttls()
            server.login(from_addr, password)
            server.sendmail(from_addr, to_addr.split(','), msg.as_string())
        return jsonify({'ok': True, 'sent': len(alerts), 'alerts': alerts})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8080))
    print(f'\n  PP/PE 레진 대시보드')
    print(f'  http://localhost:{port}\n')
    app.run(host='0.0.0.0', port=port, debug=False)
